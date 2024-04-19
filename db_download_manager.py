import os
from tqdm import tqdm
import urllib.request
import socket
import openpyxl
from openpyxl import load_workbook
from concurrent.futures import ThreadPoolExecutor, as_completed
import json
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from db_classes import GRIPdf
from pathlib import Path
import time
import csv
from dotenv import load_dotenv

# Set a default timeout for all socket operations
socket.setdefaulttimeout(5)

class DownloadManager:
    def __init__(self, folder='pdf-files', file_with_urls='GRI_2017_2020.xlsx'):
        # Create a folder to store the downloaded files
        if not os.path.exists(folder):
            os.makedirs(folder)
        self.folder = Path(__file__).parent / folder
        self.file_with_urls = file_with_urls

        load_dotenv()

        # Get the database information from the environment variables
        self.local_db_mode = os.getenv('LOCAL_DB_MODE')
        self.engine = os.getenv('ENGINE')
        self.local_db_engine = os.getenv('LOCAL_DB_ENGINE')
        self.adapter = os.getenv('ADAPTER')
        self.username = os.getenv('USERNAME')
        self.password = os.getenv('PASSWORD')
        self.hostname = os.getenv('HOSTNAME')
        self.local_db_name = os.getenv('LOCAL_DB_NAME')
        self.db_name = os.getenv('DB_NAME')
        self.db_test_name = os.getenv('DB_TEST_NAME')
    
        if self.local_db_mode == "True":
            # Use SQLite for local DB mode
            # Get the directory of this script
            dir_path = os.path.dirname(os.path.realpath(__file__))
            # Join the directory path and the local DB name to get the full path to the SQLite database file
            self.local_db_name = os.path.join(dir_path, self.local_db_name)
            DATABASE_URL = f"{self.local_db_engine}:///{self.local_db_name}"
        else:
            # Use the existing database configuration for non-local DB mode
            DATABASE_URL = f"{self.engine}+{self.adapter}://{self.username}:{self.password}@{self.hostname}/{self.db_name}"


        # Create a SQLAlchemy engine
        engine = create_engine(DATABASE_URL)

        # Create a SQLAlchemy ORM session factory
        SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
        self.engine = engine
        self.SessionLocal = SessionLocal

    def load_data(self, start=0, nrows=None):
        # Get the file extension
        _, file_extension = os.path.splitext(self.file_with_urls)
    
        # Load the file
        if file_extension == '.csv':
            with open(self.file_with_urls, 'r') as f:
                reader = csv.reader(f)
                headers = next(reader)
                for i, row in enumerate(reader):
                    if i < start:
                        continue
                    if nrows is not None and i >= start + nrows:
                        break
                    yield dict(zip(headers, row))
        elif file_extension in ['.xlsx', '.xlsm', '.xltx', '.xltm']:
            workbook = openpyxl.load_workbook(filename=self.file_with_urls, read_only=True)
            worksheet = workbook.active
            headers = [cell.value for cell in next(worksheet.iter_rows())]
            for i, row in enumerate(worksheet.iter_rows(min_row=start+2, max_row=start+nrows+1 if nrows else None), start=start):
                yield dict(zip(headers, (cell.value for cell in row)))
        else:
            raise ValueError(f"Unsupported file format: {file_extension}")
        
    def save_download_result(self, row, file_name, download_status, download_message=None, file_folder=None):
        if file_folder is None:
            file_folder = self.folder
        if download_status == 'FALSE':
            file_folder = ""
            file_name = ""
        # Convert the file_folder to a string
        file_folder = str(file_folder)
        # Start a new SQLAlchemy session
        with self.SessionLocal() as session:
            # Use the create_pdf_sync method to create a new GRIPdf object
            GRIPdf.process_row(
                session=session,
                row=row,
                file_name=file_name,
                file_folder=file_folder,
                download_status=download_status,
                download_message=download_message,
            )
            
    def download_file(self, row, url_header):
        url = row[url_header]
        brnumber = row['BRnum']
        try:
            # Generate a unique filename
            filename = f'{url.split("/")[-1]}'
    
            # If filename is None, save an error message and return 'filename_not_found'
            if filename is None:
                self.save_download_result(row, filename, download_status='FALSE', download_message='Filename could not be determined')
                return 'failed'
            
            # Check if the file already exists
            if os.path.exists(f'{self.folder}/{filename}'):
                if filename.lower().endswith('.pdf'):
                    self.save_download_result(row, filename, download_status='TRUE', download_message='File already exists in this folder')
                    return 'already_downloaded'

            # Start a new SQLAlchemy session
            with self.SessionLocal() as session:
                # Query the database for a record with the specified brnumber and download_status='TRUE'
                pdf = session.query(GRIPdf).filter_by(brnumber=brnumber, download_status='TRUE').first()

                # If such a record exists, return 'already_downloaded'
                if pdf is not None:
                    self.save_download_result(row, filename, download_status='TRUE', file_folder=pdf.file_folder, download_message=f'Download_status for {brnumber} was already TRUE, should be found in folder: {pdf.file_folder}, wont attempt download to folder: {self.folder}')
                    return 'already_downloaded'

            # Open the URL and read the first few bytes
            with urllib.request.urlopen(url, timeout=10) as u:
                if not u.read(5).startswith(b'%PDF-'):
                    self.save_download_result(row, filename, download_status='FALSE', download_message='Not received as PDF file')
                    return 'failed'
    
            # Ensure the filename ends with '.pdf'
            filename = filename if filename.endswith('.pdf') else f'{filename}.pdf'
            # Download the file
            urllib.request.urlretrieve(url, f'{self.folder}/{filename}')
            self.save_download_result(row, filename, download_status='TRUE', download_message='File downloaded successfully')
            return 'successful'
        except (socket.timeout, Exception) as e:
            self.save_download_result(row, filename, download_status='FALSE', download_message=str(e))
            return 'failed'

    def download_files(self, rows, nrows, max_workers=None):
        if max_workers is None:
            max_workers = os.cpu_count() or 1
        print(f"Attempting to download {nrows} files to folder: {self.folder} using {max_workers} logical cpu cores")
    
        # Initialize counters
        counters = {'successful': 0, 'already_downloaded': 0, 'failed': 0, 'processed_rows': 0}
    
        # Process the rows
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = {}
            for row in rows:
                # Submit the Pdf_URL download task to the executor
                future = executor.submit(self.download_file, row, 'Pdf_URL')
                futures[future] = (row, 'Pdf_URL')
    
            # Wait for tasks to complete and update counters
            for future in as_completed(futures):
                result = future.result()
                if result in counters:
                    counters[result] += 1
    
                row, url_header = futures[future]
                if result == 'failed' and url_header == 'Pdf_URL':
                    # If the Pdf_URL download task failed, submit the Report Html Address download task to the executor
                    future = executor.submit(self.download_file, row, 'Report Html Address')
                    futures[future] = (row, 'Report Html Address')
    
                # Increment the processed_rows counter
                counters['processed_rows'] += 1
    
                # Yield the current counters
                yield counters

    def start_download(self, start_row, nrows):
        start_time = time.time()
        rows = self.load_data(start=start_row, nrows=nrows)
    
        # Initialize counters
        counters = {'successful': 0, 'already_downloaded': 0, 'failed': 0}
    
        # Process the rows
        with tqdm(total=nrows) as pbar:
            for result in self.download_files(rows, nrows):
                counters.update(result)
    
                # Calculate the elapsed time
                elapsed_time = time.time() - start_time
                m, s = divmod(elapsed_time, 60)
                h, m = divmod(m, 60)
    
                # Update the description of the progress bar with the status information
                pbar.set_description(f"Elapsed time: {int(h)} hours, {int(m)} minutes, {s:.2f} seconds, "
                                     f"Success: {counters['successful']}, "
                                     f"Existed: {counters['already_downloaded']}, "
                                     f"Failed: {counters['failed']}")
    
                pbar.update()  # Update the progress bar

                yield counters

def main():
    dm = DownloadManager(folder='pdf-files', file_with_urls='pdf-urls/GRI_2017_2020.xlsx')
    start_row = 0
    nrows = 20
    for result in dm.start_download(start_row, nrows):
        pass

if __name__ == '__main__':
    main()