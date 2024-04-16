import os
from tqdm import tqdm
import urllib.request
import socket
from openpyxl import load_workbook
from concurrent.futures import ThreadPoolExecutor, as_completed
import json
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from db_classes import GRIPdf

# Set a default timeout for all socket operations
socket.setdefaulttimeout(5)

class DownloadManager:
    def __init__(self, folder='pdf-files', file_with_urls='GRI_2017_2020.xlsx'):
        # Create a folder to store the downloaded files
        if not os.path.exists(folder):
            os.makedirs(folder)
        self.folder = folder
        self.file_with_urls = file_with_urls

        # Load database information from db_info.json
        with open('db_info.json') as f:
            db_info = json.load(f)

        # Define the database URL
        DATABASE_URL = f"mysql+pymysql://{db_info['username']}:{db_info['password']}@{db_info['hostname']}/{db_info['db_name']}"

        # Create a SQLAlchemy engine
        engine = create_engine(DATABASE_URL)

        # Create a SQLAlchemy ORM session factory
        SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
        self.engine = engine
        self.SessionLocal = SessionLocal

    def load_data(self, start=0, nrows=None):
        # Load the workbook
        workbook = load_workbook(filename=self.file_with_urls, read_only=True)

        # Get the first worksheet
        worksheet = workbook.active

        # Get the first row (column headers)
        headers = [cell.value for cell in next(worksheet.iter_rows())]

        # Yield each row in the range
        for i, row in enumerate(worksheet.iter_rows(min_row=start+2, max_row=start+nrows+1 if nrows else None), start=start):  # start+2 because 1-based index and header row
            # Yield the row as a dictionary
            yield dict(zip(headers, (cell.value for cell in row)))

    def save_download_result(self, row, file_name, download_status, download_message=None, file_folder=None):
        if file_folder is None:
            file_folder = self.folder
        if download_status == 'FALSE':
            file_folder = ""
            file_name = ""
        # Start a new SQLAlchemy session
        with self.SessionLocal() as session:
            # Use the create_pdf_sync method to create a new GRIPdf object
            GRIPdf.process_row(
                session=session,
                row=row,
                file_name=file_name,
                file_folder=file_folder,
                download_status=download_status,
                download_message=download_message,
            )
            
    def download_file(self, row, url_header):
        url = row[url_header]
        brnumber = row['BRnum']
        try:
            # Generate a unique filename
            filename = f'{url.split("/")[-1]}'
    
            # If filename is None, save an error message and return 'filename_not_found'
            if filename is None:
                self.save_download_result(row, filename, download_status='FALSE', download_message='Filename could not be determined')
                return 'failed'
            
            # Check if the file already exists
            if os.path.exists(f'{self.folder}/{filename}'):
                if filename.lower().endswith('.pdf'):
                    self.save_download_result(row, filename, download_status='TRUE', download_message='File already exists in this folder')
                    return 'already_downloaded'

            # Start a new SQLAlchemy session
            with self.SessionLocal() as session:
                # Query the database for a record with the specified brnumber and download_status='TRUE'
                pdf = session.query(GRIPdf).filter_by(brnumber=brnumber, download_status='TRUE').first()

                # If such a record exists, return 'already_downloaded'
                if pdf is not None:
                    self.save_download_result(row, filename, download_status='TRUE', file_folder=pdf.file_folder, download_message=f'Download_status for {brnumber} was already TRUE, should be found in folder: {pdf.file_folder}, wont attempt download to folder: {self.folder}')
                    return 'already_downloaded'

            # Open the URL and read the first few bytes
            with urllib.request.urlopen(url, timeout=10) as u:
                if not u.read(5).startswith(b'%PDF-'):
                    self.save_download_result(row, filename, download_status='FALSE', download_message='Not received as PDF file')
                    return 'failed'
    
            # Download the file
            urllib.request.urlretrieve(url, f'{self.folder}/{filename}')
            self.save_download_result(row, filename, download_status='TRUE', download_message='File downloaded successfully')
            return 'successful'
        except (socket.timeout, Exception) as e:
            self.save_download_result(row, filename, download_status='FALSE', download_message=str(e))
            return 'failed'

    def download_files(self, rows, nrows, max_workers=None):
        if max_workers is None:
            max_workers = os.cpu_count() or 1
        print(f"Attempting to download {nrows} files to folder: {self.folder} using {max_workers} logical cpu cores")
    
        # Initialize counters
        counters = {'successful': 0, 'already_downloaded': 0}
    
        # Process the rows
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = {}
            for row in tqdm(rows, desc="Processing data", total=nrows):
                # Submit the Pdf_URL download task to the executor
                future = executor.submit(self.download_file, row, 'Pdf_URL')
                futures[future] = (row, 'Pdf_URL')

            # Wait for tasks to complete and update counters
            for future in tqdm(as_completed(futures), total=len(futures), desc="Downloading files"):
                result = future.result()
                if result in counters:
                    counters[result] += 1

                row, url_header = futures[future]
                if result == 'failed' and url_header == 'Pdf_URL':
                    # If the Pdf_URL download task failed, submit the Report Html Address download task to the executor
                    future = executor.submit(self.download_file, row, 'Report Html Address')
                    futures[future] = (row, 'Report Html Address')

        # Calculate the number of failed downloads
        counters['failed'] = nrows - counters['successful'] - counters['already_downloaded']

        return counters['successful'], counters['failed'], counters['already_downloaded']

    def start_download(self, start_row, nrows):
        rows = self.load_data(start=start_row, nrows=nrows)
        successful_downloads, failed_downloads, already_downloaded = self.download_files(rows, nrows)
        print(f"Successfully downloaded: {successful_downloads}")
        print(f"Already downloaded: {already_downloaded}")
        print(f"Failed to download: {failed_downloads}")

def main():
    dm = DownloadManager(folder='pdf-files', file_with_urls='GRI_2017_2020.xlsx')
    dm.start_download(start_row=125, nrows=50)

if __name__ == '__main__':
    main()