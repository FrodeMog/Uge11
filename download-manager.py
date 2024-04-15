import os
import pandas as pd
import requests
from tqdm import tqdm
import urllib.request
import socket
from openpyxl import load_workbook
from concurrent.futures import ThreadPoolExecutor, as_completed

# Set a default timeout for all socket operations
socket.setdefaulttimeout(5)

def load_data(file, start=0, nrows=None):
    # Load the workbook
    workbook = load_workbook(filename=file, read_only=True)

    # Get the first worksheet
    worksheet = workbook.active

    # Get the first row (column headers)
    headers = [cell.value for cell in next(worksheet.iter_rows())]

    # Yield each row in the range
    for i, row in enumerate(worksheet.iter_rows(min_row=start+2, max_row=start+nrows+1 if nrows else None), start=start):  # start+2 because 1-based index and header row
        # Yield the row as a dictionary
        yield dict(zip(headers, (cell.value for cell in row)))

def download_file(url, folder):
    try:
        # Generate a unique filename
        filename = f'{url.split("/")[-1]}'

        # Check if the file already exists
        if os.path.exists(f'{folder}/{filename}'):
            if filename.lower().endswith('.pdf'):
                return 'already_downloaded'

        # Check if the file is a .pdf
        if not filename.lower().endswith('.pdf'):
            return 'failed'

        # Open the URL and read the first few bytes
        with urllib.request.urlopen(url, timeout=10) as u:
            if not u.read(5).startswith(b'%PDF-'):
                return 'failed'

        # Download the file
        urllib.request.urlretrieve(url, f'{folder}/{filename}')
        return 'successful'
    except (socket.timeout, Exception) as e:
        return 'failed'

def download_files(rows, nrows, folder='pdf-files', max_workers=5):
    # Create a folder to store the downloaded files
    if not os.path.exists(folder):
        os.makedirs(folder)

    # Initialize counters
    counters = {'successful': 0, 'already_downloaded': 0}

    # Process the rows
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        # Submit all tasks to the executor
        futures = {executor.submit(download_file, url, folder): url for row in tqdm(rows, desc="Processing files", total=nrows) for url in [row['Pdf_URL'], row['Report Html Address']]}

        # Wait for tasks to complete and update counters
        for future in tqdm(as_completed(futures), total=len(futures), desc="Downloading files"):
            result = future.result()
            if result in counters:
                counters[result] += 1

    # Calculate the number of failed downloads
    counters['failed'] = nrows - counters['successful'] - counters['already_downloaded']

    return counters['successful'], counters['failed'], counters['already_downloaded']

def main():
    start = 0
    nrows = 20
    print("Loading rows")
    rows = load_data(file='GRI_2017_2020.xlsx', start=start, nrows=nrows)
    successful_downloads, failed_downloads, already_downloaded = download_files(rows, nrows, folder='pdf-files')
    print(f"Successfully downloaded: {successful_downloads}")
    print(f"Already downloaded: {already_downloaded}")
    print(f"Failed to download: {failed_downloads}")

if __name__ == '__main__':
    main()